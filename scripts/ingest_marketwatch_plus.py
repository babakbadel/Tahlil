"""Ingest the repository's MarketWatchPlus Excel snapshot into BabiMind.

The workbook schema can change between market exports, so this loader keeps the
original sheet/header structure while producing a normalized JSON representation.
It never fabricates missing market fields.
"""
from __future__ import annotations

import json
import math
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "data/latest/MarketWatchPlus-1405_06_07.xlsx"
OUTPUT = ROOT / "data/raw/marketwatch_plus_snapshot.json"
REPORT = ROOT / "reports/marketwatch_plus_summary.json"


def clean(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
        return None
    if isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def main() -> None:
    if not SOURCE.exists():
        raise FileNotFoundError(SOURCE)

    wb = load_workbook(SOURCE, read_only=True, data_only=True)
    sheets: dict[str, Any] = {}
    total_rows = 0
    total_data_rows = 0

    for ws in wb.worksheets:
        rows = ws.iter_rows(values_only=True)
        try:
            first = next(rows)
        except StopIteration:
            sheets[ws.title] = {"headers": [], "records": [], "row_count": 0}
            continue

        headers = [clean(v) for v in first]
        records: list[dict[str, Any]] = []
        row_count = 0
        for raw in rows:
            row_count += 1
            values = [clean(v) for v in raw]
            # Preserve duplicate/blank Excel headers without silently losing columns.
            keys: list[str] = []
            seen: dict[str, int] = {}
            for i, h in enumerate(headers):
                base = str(h).strip() if h not in (None, "") else f"column_{i+1}"
                seen[base] = seen.get(base, 0) + 1
                keys.append(base if seen[base] == 1 else f"{base}_{seen[base]}")
            records.append({k: values[i] if i < len(values) else None for i, k in enumerate(keys)})

        sheets[ws.title] = {
            "headers": headers,
            "row_count": row_count,
            "records": records,
        }
        total_rows += row_count + 1
        total_data_rows += row_count

    wb.close()
    payload = {
        "source": str(SOURCE.relative_to(ROOT)),
        "source_type": "repository_excel_snapshot",
        "snapshot_date": "1405-06-07",
        "generated_at": datetime.utcnow().isoformat() + "Z",
        "sheet_count": len(sheets),
        "total_data_rows": total_data_rows,
        "total_rows_including_headers": total_rows,
        "sheets": sheets,
        "data_quality": {
            "schema_preserved": True,
            "missing_values_preserved": True,
            "greeks_or_iv_inferred": False,
            "warning": "Workbook values are ingested as published. No IV, Greeks, timestamps or market semantics are inferred when absent."
        },
    }

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    REPORT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2, default=str), encoding="utf-8")

    summary = {
        "source": payload["source"],
        "snapshot_date": payload["snapshot_date"],
        "generated_at": payload["generated_at"],
        "sheet_count": payload["sheet_count"],
        "total_data_rows": payload["total_data_rows"],
        "sheets": {
            name: {"row_count": info["row_count"], "headers": info["headers"]}
            for name, info in sheets.items()
        },
        "data_quality": payload["data_quality"],
    }
    REPORT.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"MarketWatchPlus: {len(sheets)} sheets / {total_data_rows} data rows -> {OUTPUT}")


if __name__ == "__main__":
    main()
